import requests
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import os
from dotenv import load_dotenv


load_dotenv()
FMP_API_KEY = os.getenv("FMP_API_KEY")

def user_input():


    symbols = []
    while True:
        symbol = input("Enter stock symbol (uppercase) (or 'done' to finish): ").upper()
        if symbol == "DONE":
            if not symbols:
                print("Please enter at least one symbol.")
            else:
                break
        else:
            symbols.append(symbol)
    return symbols


def get_income_statement(api_key, symbol, period):


    url = f"https://financialmodelingprep.com/api/v3/income-statement/{symbol}?period=annual&apikey={api_key}"
    response = requests.get(url)
    print(symbol , "got statement")
    return pd.DataFrame(response.json())



def get_earnings_data(api_key, start_date, end_date):

    url = f"https://financialmodelingprep.com/api/v3/earning_calendar?from={start_date}&to={end_date}&apikey={api_key}"
    print(url)
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        return pd.DataFrame(data)
    else:
        print(f"Failed to fetch earnings data from FinancialModelingPrep API. Status code: {response.status_code}")
        return None
    

def get_market_cap(api_key, symbol):
    url = f"https://financialmodelingprep.com/api/v3/market-capitalization/{symbol}?apikey={api_key}"
    response = requests.get(url)
    if response.status_code == 200:
        marketCap = ""
        data = response.json()
        if len(data) > 0:
            marketCap = data[0].get("marketCap")
        else:
            marketCap = "NA"
        return marketCap
    else:
        print(f"Failed to fetch market capitalization from FinancialModelingPrep API. Status code: {response.status_code}")
        return None
    
def create_excel_report(data, symbols):
  
    if data is None:
        print("No data available for creating Excel report.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Earnings Report"

    header_row = 1
    headers = ["Stock Ticker", "Date", "EPS", "Revenue", "EPS Growth YoY", "Revenue Growth YoY", "Market Cap", "Short Interest"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=idx).value = header

    data_row = 2
    for symbol in symbols:
        filtered_data = data[data['symbol'] == symbol]  # Filter data for the symbol
        if len(filtered_data) == 0:
            print(f"No data available for symbol {symbol}.")
            continue

        income_statement_data = get_income_statement(FMP_API_KEY, symbol, 'annual')
        if income_statement_data is None or len(income_statement_data) < 2:
            print(f"Insufficient income statement data for YoY calculations for {symbol}.")
            continue
        
        market_cap = get_market_cap(FMP_API_KEY, symbol)
        if market_cap is None:
            market_cap = "NA"
        
        current_year_data = income_statement_data.iloc[0]
        previous_year_data = income_statement_data.iloc[1]
        current_eps = current_year_data["eps"]
        previous_eps = previous_year_data["eps"]
        eps_yoy_growth = ((current_eps - previous_eps) / previous_eps) * 100

        current_revenue = current_year_data["revenue"]
        previous_revenue = previous_year_data["revenue"]
        revenue_yoy_growth = ((current_revenue - previous_revenue) / previous_revenue) * 100

        for index, row in filtered_data.iterrows():
            ws.cell(row=data_row, column=1).value = symbol  # Stock ticker
            ws.cell(row=data_row, column=2).value = row["date"]  # Date
            ws.cell(row=data_row, column=3).value = row.get("eps", "N/A")  # EPS
            ws.cell(row=data_row, column=4).value = row.get("revenue", "N/A")  # Revenue
            ws.cell(row=data_row, column=5).value = f"{eps_yoy_growth:.2f}%"  # EPS Growth YoY
            ws.cell(row=data_row, column=6).value = f"{revenue_yoy_growth:.2f}%"  # Revenue Growth YoY
            ws.cell(row=data_row, column=7).value = {market_cap}  # Market Cap (data from another source)
            # ws.cell(row=data_row, column=8).value = ""  # Short Interest (data from another source)
            data_row += 1

    for col in ws.iter_cols(min_col=3, max_col=6):
        for cell in col:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'


    file_name = f"Earnings_report_{datetime.today().strftime('%Y-%m-%d')}.xlsx"
    wb.save(file_name)
    print(f"Earnings report saved as {file_name}.")

def main():
    today = datetime.today()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    start_date = start_of_week.strftime("%Y-%m-%d")
    end_date = end_of_week.strftime("%Y-%m-%d")
    earnings_data = get_earnings_data(FMP_API_KEY, start_date, end_date)

    if earnings_data is None:
        print("Exiting program.")
        return

    symbols = user_input()
    if not symbols:
        print("No stock symbols entered. Exiting program.")
        return

    create_excel_report(earnings_data.copy(), symbols)

    print("Earnings report generation complete.")

if __name__ == "__main__":
    main()
